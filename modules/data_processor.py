"""
데이터 정규화, 집계 및 엑셀(.xlsx) 생성 모듈 (Data Processor & Excel Exporter)
"""
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Tuple
from config import STANDARD_COLUMNS

class DataProcessor:
    """
    추출된 레코드 목록을 정규화, 기업별 합산, 필터링 및 엑셀 파일로 변환합니다.
    """

    def __init__(self):
        self.columns = STANDARD_COLUMNS

    def create_dataframe(self, records: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        추출 레코드 리스트를 표준 DataFrame으로 변환합니다.
        """
        if not records:
            return pd.DataFrame(columns=self.columns)

        data = []
        for r in records:
            data.append({
                "나라": r.get("country", "기타"),
                "기업": r.get("company", "(주)미상기업"),
                "성과월": r.get("month", "2024-01"),
                "수출액": float(r.get("amount", 0.0)),
                "문서유형": r.get("doc_type", "미확인"),
                "파일명": r.get("file_name", "")
            })

        df = pd.DataFrame(data)
        # 수출액 내림차순 정렬
        df = df.sort_values(by=["기업", "성과월"], ascending=[True, True]).reset_index(drop=True)
        return df

    def get_company_summary(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        기업별 총수출 성약액 및 건수를 집계합니다.
        """
        if df.empty:
            return pd.DataFrame(columns=["기업", "총수출 성약액", "건수"])

        summary = df.groupby("기업").agg(
            총수출_성약액=("수출액", "sum"),
            건수=("수출액", "count")
        ).reset_index()

        summary.rename(columns={"총수출_성약액": "총수출 성약액"}, inplace=True)
        summary = summary.sort_values(by="총수출 성약액", ascending=False).reset_index(drop=True)
        return summary

    def filter_by_company(self, df: pd.DataFrame, company_name: str) -> pd.DataFrame:
        """
        선택된 기업명으로 데이터를 필터링하고 [나라, 기업, 성과월, 수출액] 표준 컬럼만 반환합니다.
        """
        if df.empty:
            return pd.DataFrame(columns=self.columns)

        if company_name == "전체" or not company_name:
            filtered = df[self.columns].copy()
        else:
            filtered = df[df["기업"] == company_name][self.columns].copy()

        return filtered.reset_index(drop=True)

    def export_to_excel_bytes(self, df: pd.DataFrame, company_title: str = "전체") -> bytes:
        """
        데이터프레임을 서식이 적용된 표준 엑셀 파일(.xlsx) 바이너리로 내보냅니다.
        """
        # 출력 대상 컬럼만 선택
        export_df = df[self.columns].copy() if all(c in df.columns for c in self.columns) else df.copy()

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            sheet_name = f"{company_title[:25]} 수출내역"
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            worksheet = writer.sheets[sheet_name]

            # 스타일 정의
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
            data_font = Font(name="맑은 고딕", size=10)
            center_align = Alignment(horizontal="center", vertical="center")
            right_align = Alignment(horizontal="right", vertical="center")
            
            thin_border = Border(
                left=Side(style="thin", color="D3D3D3"),
                right=Side(style="thin", color="D3D3D3"),
                top=Side(style="thin", color="D3D3D3"),
                bottom=Side(style="thin", color="D3D3D3")
            )

            # 1. 헤더 행 서식 지정
            for col_num in range(1, len(export_df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            # 2. 데이터 행 서식 및 숫자 콤마 지정
            for row_num in range(2, len(export_df) + 2):
                for col_num in range(1, len(export_df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = data_font
                    cell.border = thin_border

                    col_name = export_df.columns[col_num - 1]
                    if col_name == "수출액":
                        cell.number_format = "#,##0"
                        cell.alignment = right_align
                    else:
                        cell.alignment = center_align

            # 3. 컬럼 너비 자동 맞춤
            for col in worksheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                worksheet.column_dimensions[col_letter].width = max(max_len + 5, 14)

        output.seek(0)
        return output.getvalue()
